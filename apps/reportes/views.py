from datetime import date, datetime, timedelta, time as time_type
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from .models import PlantillaReporte
from apps.asistencia.models import AsistenciaDiaria
from apps.asistencia.calculators.engine import obtener_horario_empleado
from apps.empleados.models import Empleado


@login_required
def lista_reportes(request):
    return render(request, 'reportes/lista.html')


@login_required
@staff_member_required
def generar_reporte(request):
    tipo = request.GET.get('tipo', 'diario')
    formato = request.GET.get('formato', 'xlsx')
    desde = request.GET.get('desde', str(date.today()))
    hasta = request.GET.get('hasta', str(date.today()))

    if tipo == 'secretaria':
        if formato == 'xlsx':
            return _exportar_secretaria_excel(desde, hasta)
        elif formato == 'csv':
            return _exportar_secretaria_csv(desde, hasta)
        elif formato == 'pdf':
            return _exportar_secretaria_pdf(desde, hasta)
        messages.error(request, 'Formato no soportado')
        return redirect('lista-reportes')

    if formato == 'xlsx':
        return _exportar_excel(tipo, desde, hasta)
    elif formato == 'csv':
        return _exportar_csv(tipo, desde, hasta)
    elif formato == 'pdf':
        return _exportar_pdf(tipo, desde, hasta)

    messages.error(request, 'Formato no soportado')
    return redirect('lista-reportes')


@login_required
def lista_plantillas(request):
    plantillas = PlantillaReporte.objects.all()
    return render(request, 'reportes/plantillas.html', {
        'plantillas': plantillas,
    })


def _exportar_excel(tipo, desde, hasta):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    qs = _obtener_datos(tipo, desde, hasta)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte'

    headers = ['ID', 'Nombre', 'Departamento', 'Fecha', 'Entrada',
               'Comida Inicio', 'Comida Fin', 'Salida', 'Horas',
               'Retardo (min)', 'Incidencia', 'Estatus']

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for i, a in enumerate(qs, 2):
        ws.cell(row=i, column=1, value=a.empleado.id_original)
        ws.cell(row=i, column=2, value=a.empleado.nombre)
        ws.cell(row=i, column=3, value=str(a.empleado.departamento or ''))
        ws.cell(row=i, column=4, value=a.fecha.isoformat())
        ws.cell(row=i, column=5, value=str(a.entrada or ''))
        ws.cell(row=i, column=6, value=str(a.comida_inicio or ''))
        ws.cell(row=i, column=7, value=str(a.comida_fin or ''))
        ws.cell(row=i, column=8, value=str(a.salida or ''))
        ws.cell(row=i, column=9, value=float(a.horas_jornada))
        ws.cell(row=i, column=10, value=a.minutos_retardo)
        ws.cell(row=i, column=11, value=a.incidencia_codigo)
        ws.cell(row=i, column=12, value=a.estatus)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=reporte_{tipo}_{date.today()}.xlsx'
    wb.save(response)
    return response


def _exportar_csv(tipo, desde, hasta):
    import csv
    import io

    qs = _obtener_datos(tipo, desde, hasta)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['ID', 'Nombre', 'Departamento', 'Fecha', 'Entrada',
                     'Comida Inicio', 'Comida Fin', 'Salida', 'Horas',
                     'Retardo (min)', 'Incidencia', 'Estatus'])

    for a in qs:
        writer.writerow([
            a.empleado.id_original, a.empleado.nombre,
            str(a.empleado.departamento or ''), a.fecha.isoformat(),
            str(a.entrada or ''), str(a.comida_inicio or ''),
            str(a.comida_fin or ''), str(a.salida or ''),
            float(a.horas_jornada), a.minutos_retardo,
            a.incidencia_codigo, a.estatus,
        ])

    response = HttpResponse(buffer.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename=reporte_{tipo}_{date.today()}.csv'
    return response


def _exportar_pdf(tipo, desde, hasta):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    import io

    qs = _obtener_datos(tipo, desde, hasta)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []

    styles = getSampleStyleSheet()
    elements.append(Paragraph(f'Reporte de {tipo}', styles['Title']))

    data = [['ID', 'Nombre', 'Depto', 'Fecha', 'Entrada', 'Salida', 'Horas', 'Retardo', 'Incidencia']]
    for a in qs:
        data.append([
            a.empleado.id_original,
            a.empleado.nombre[:20],
            str(a.empleado.departamento or '')[:15],
            a.fecha.isoformat(),
            str(a.entrada or ''),
            str(a.salida or ''),
            f'{float(a.horas_jornada):.2f}',
            str(a.minutos_retardo),
            a.incidencia_codigo or '-',
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    elements.append(table)
    doc.build(elements)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=reporte_{tipo}_{date.today()}.pdf'
    return response


def _obtener_datos(tipo, desde, hasta):
    qs = AsistenciaDiaria.objects.select_related('empleado__departamento')

    if desde:
        qs = qs.filter(fecha__gte=desde)
    if hasta:
        qs = qs.filter(fecha__lte=hasta)

    if tipo == 'retardos':
        qs = qs.filter(minutos_retardo__gt=0)
    elif tipo == 'incidencias':
        qs = qs.exclude(incidencia_codigo='')
    elif tipo == 'ausentes':
        qs = qs.filter(estatus='ausente')

    return qs.order_by('empleado__nombre', 'fecha')


def _obtener_datos_secretaria(desde, hasta):
    empleados = Empleado.objects.filter(estatus='activo').order_by('nombre')
    desde_date = date.fromisoformat(desde)
    hasta_date = date.fromisoformat(hasta)
    delta = hasta_date - desde_date
    filas = []

    for emp in empleados:
        for i in range(delta.days + 1):
            dia = desde_date + timedelta(days=i)
            horario_info, es_excepcion = obtener_horario_empleado(emp, dia)
            if horario_info is None:
                continue

            if es_excepcion:
                entrada_time = horario_info.entrada_esperada
                salida_programada = horario_info.salida_esperada
                if entrada_time and salida_programada:
                    diff = datetime.combine(dia, salida_programada) - datetime.combine(dia, entrada_time)
                    jornada_hrs = round(diff.total_seconds() / 3600, 2)
                else:
                    jornada_hrs = 0
                entrada = entrada_time
                salida_fija = salida_programada
            else:
                horario = horario_info
                entrada_base = horario.ventana_entrada_inicio
                jornada_hrs = float(horario.jornada_hrs)
                asist = AsistenciaDiaria.objects.filter(empleado=emp, fecha=dia).first()
                entrada = asist.entrada if asist and asist.entrada else entrada_base
                h = int(jornada_hrs)
                m = int((jornada_hrs - h) * 60)
                salida_dt = datetime.combine(dia, entrada) + timedelta(hours=h, minutes=m)
                salida_fija = salida_dt.time()

            filas.append({
                'empleado': emp,
                'fecha': dia,
                'entrada': entrada,
                'salida_fija': salida_fija,
                'jornada': jornada_hrs,
            })

    return filas


def _exportar_secretaria_excel(desde, hasta):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    filas = _obtener_datos_secretaria(desde, hasta)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Horas Secretaría'

    headers = ['ID', 'Nombre', 'Departamento', 'Fecha', 'Hora Entrada', 'Hora Salida (8h)', 'Jornada']
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for i, f in enumerate(filas, 2):
        ws.cell(row=i, column=1, value=f['empleado'].id_original)
        ws.cell(row=i, column=2, value=f['empleado'].nombre)
        ws.cell(row=i, column=3, value=str(f['empleado'].departamento or ''))
        ws.cell(row=i, column=4, value=f['fecha'].isoformat())
        ws.cell(row=i, column=5, value=str(f['entrada'] or ''))
        ws.cell(row=i, column=6, value=str(f['salida_fija'] or ''))
        ws.cell(row=i, column=7, value=float(f['jornada']))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=horas_secretaria_{date.today()}.xlsx'
    wb.save(response)
    return response


def _exportar_secretaria_csv(desde, hasta):
    import csv
    import io

    filas = _obtener_datos_secretaria(desde, hasta)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['ID', 'Nombre', 'Departamento', 'Fecha', 'Hora Entrada', 'Hora Salida (8h)', 'Jornada'])

    for f in filas:
        writer.writerow([
            f['empleado'].id_original,
            f['empleado'].nombre,
            str(f['empleado'].departamento or ''),
            f['fecha'].isoformat(),
            str(f['entrada'] or ''),
            str(f['salida_fija'] or ''),
            float(f['jornada']),
        ])

    response = HttpResponse(buffer.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename=horas_secretaria_{date.today()}.csv'
    return response


def _exportar_secretaria_pdf(desde, hasta):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    import io

    filas = _obtener_datos_secretaria(desde, hasta)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []

    styles = getSampleStyleSheet()
    elements.append(Paragraph('Horas Secretaría del Trabajo', styles['Title']))

    data = [['ID', 'Nombre', 'Depto', 'Fecha', 'Entrada', 'Salida (8h)', 'Jornada']]
    for f in filas:
        data.append([
            f['empleado'].id_original,
            f['empleado'].nombre[:20],
            str(f['empleado'].departamento or '')[:15],
            f['fecha'].isoformat(),
            str(f['entrada'] or ''),
            str(f['salida_fija'] or ''),
            f'{float(f["jornada"]):.2f}',
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    elements.append(table)
    doc.build(elements)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=horas_secretaria_{date.today()}.pdf'
    return response
