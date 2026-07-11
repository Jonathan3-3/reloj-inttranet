from datetime import date, datetime, timedelta
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from django.db.models import Q
from .models import Marcacion, AsistenciaDiaria
from apps.empleados.models import Empleado
from .calculators.engine import recalcular_asistencia, obtener_horario_empleado


@login_required
def buscar_asistencia(request):
    return render(request, 'asistencia/buscar.html', {
        'hoy': timezone.localtime().date(),
    })


@login_required
def vista_hoy(request):
    hoy = timezone.localtime().date()
    registros_hoy = Marcacion.objects.filter(marcado_en__date=hoy).count()
    total_empleados = Empleado.objects.filter(estatus='activo').count()
    asistencias_hoy = AsistenciaDiaria.objects.filter(fecha=hoy)
    asistencias_count = asistencias_hoy.count()
    retardos_hoy = asistencias_hoy.filter(incidencia_codigo='llt').count()

    return render(request, 'asistencia/hoy.html', {
        'hoy': hoy,
        'total_empleados': total_empleados,
        'registros_hoy': registros_hoy,
        'asistencias_count': asistencias_count,
        'retardos_hoy': retardos_hoy,
    })


@login_required
def detalle_asistencia(request, pk):
    empleado = get_object_or_404(Empleado, pk=pk)
    return render(request, 'asistencia/detalle.html', {
        'empleado': empleado,
    })


# ─── API ───

@login_required
def api_reporte(request):
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')
    empleado_id = request.GET.get('empleado_id')
    nombre = request.GET.get('nombre', '')

    qs = AsistenciaDiaria.objects.select_related('empleado__departamento', 'horario')

    if desde:
        qs = qs.filter(fecha__gte=desde)
    if hasta:
        qs = qs.filter(fecha__lte=hasta)
    if empleado_id:
        qs = qs.filter(empleado_id=empleado_id)
    if nombre:
        qs = qs.filter(empleado__nombre__icontains=nombre)

    qs = qs.order_by('empleado__nombre', 'fecha')

    resultados = []
    for a in qs:
        resultados.append({
            'id': a.empleado.id_original,
            'nombre': a.empleado.nombre,
            'departamento': str(a.empleado.departamento or ''),
            'fecha': a.fecha.isoformat(),
            'entrada': str(a.entrada or ''),
            'comida_inicio': str(a.comida_inicio or ''),
            'comida_fin': str(a.comida_fin or ''),
            'salida': str(a.salida or ''),
            'horas': float(a.horas_jornada),
            'retardo': a.minutos_retardo,
            'comida_min': a.minutos_comida,
            'comida_excedida': a.comida_excedida,
            'horas_extra_minutos': a.horas_extra_minutos,
            'incidencia': a.incidencia_codigo,
            'estatus': a.estatus,
        })

    return JsonResponse({
        'resultados': resultados,
        'total': len(resultados),
    })


@login_required
def api_reporte_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')
    empleado_pk = request.GET.get('empleado_pk')

    qs = AsistenciaDiaria.objects.select_related('empleado__departamento')

    if desde:
        qs = qs.filter(fecha__gte=desde)
    if hasta:
        qs = qs.filter(fecha__lte=hasta)
    if empleado_pk:
        qs = qs.filter(empleado_id=empleado_pk)

    qs = qs.order_by('empleado__nombre', 'fecha')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Detalle'

    headers = ['ID', 'Nombre', 'Departamento', 'Fecha', 'Entrada',
               'Comida Inicio', 'Comida Fin', 'Salida', 'Horas',
               'Retardo (min)', 'Comida Excedida', 'Horas Extra (min)',
               'Incidencia', 'Estatus']

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    incidencia_fills = {
        'llt': PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),
        'finj': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),
    }
    justified_fill = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')

    for i, a in enumerate(qs, 2):
        ws.cell(row=i, column=1, value=a.empleado.id_original)
        ws.cell(row=i, column=2, value=a.empleado.nombre)
        ws.cell(row=i, column=3, value=str(a.empleado.departamento or ''))
        ws.cell(row=i, column=4, value=a.fecha.isoformat())
        ws.cell(row=i, column=5, value=str(a.entrada or ''))
        ws.cell(row=i, column=6, value=str(a.comida_inicio or ''))
        ws.cell(row=i, column=7, value=str(a.comida_fin or ''))
        ws.cell(row=i, column=8, value=str(a.salida or ''))
        ws.cell(row=i, column=9, value=float(a.horas_jornada))
        ws.cell(row=i, column=10, value=a.minutos_retardo)
        ws.cell(row=i, column=11, value='Sí' if a.comida_excedida else 'No')
        ws.cell(row=i, column=12, value=a.horas_extra_minutos)
        ws.cell(row=i, column=13, value=a.incidencia_codigo)
        ws.cell(row=i, column=14, value=a.estatus)

        if a.incidencia_codigo in incidencia_fills:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = incidencia_fills[a.incidencia_codigo]
        elif a.incidencia_codigo:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = justified_fill

    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(qs) + 1}'

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=reporte_asistencia_{date.today()}.xlsx'
    wb.save(response)
    return response


@login_required
def api_today(request):
    hoy = timezone.localtime().date()
    hace_5min = timezone.now() - timedelta(minutes=5)

    recientes = Marcacion.objects.filter(
        marcado_en__date=hoy
    ).select_related('empleado').order_by('-marcado_en')[:20]

    conexiones = []
    try:
        from apps.registro.models import ConexionWeb
        conns = ConexionWeb.objects.filter(
            activa=True, ultimo_ping__gte=hace_5min
        ).select_related('empleado')
        for c in conns:
            conexiones.append({
                'id': c.empleado.id_original,
                'nombre': c.empleado.nombre,
                'tipo': c.tipo_dispositivo,
                'ip': c.ip_address,
                'tiempo': c.tiempo_conectado,
            })
    except Exception:
        pass

    return JsonResponse({
        'registros': [{
            'id': r.empleado.id_original,
            'nombre': r.empleado.nombre,
            'hora': timezone.localtime(r.marcado_en).strftime('%H:%M:%S'),
            'fuente': r.fuente,
        } for r in recientes],
        'conexiones': conexiones,
        'total_registros_hoy': Marcacion.objects.filter(marcado_en__date=hoy).count(),
    })


@login_required
def api_empleados(request):
    q = request.GET.get('q', '')
    empleados = Empleado.objects.filter(estatus='activo')
    if q:
        empleados = empleados.filter(
            Q(id_original__icontains=q) | Q(nombre__icontains=q)
        )
    empleados = empleados.order_by('nombre')[:30]

    return JsonResponse({
        'empleados': [{
            'id': e.id,
            'id_original': e.id_original,
            'nombre': e.nombre,
            'departamento': str(e.departamento or ''),
        } for e in empleados]
    })


@login_required
def reporte_horas_reales(request):
    empleados = Empleado.objects.filter(estatus='activo').order_by('nombre')
    resultados = []
    empleado_id = request.GET.get('empleado_id', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')

    if empleado_id and desde and hasta:
        qs = AsistenciaDiaria.objects.filter(
            empleado_id=empleado_id,
            fecha__gte=desde,
            fecha__lte=hasta,
        ).select_related('empleado', 'horario').order_by('fecha')

        for a in qs:
            resultados.append({
                'fecha': a.fecha,
                'entrada': a.entrada,
                'comida_inicio': a.comida_inicio,
                'comida_fin': a.comida_fin,
                'salida': a.salida,
                'horas': a.horas_jornada,
                'retardo': a.minutos_retardo,
                'comida_excedida': a.comida_excedida,
                'horas_extra_minutos': a.horas_extra_minutos,
                'incidencia': a.incidencia_codigo,
                'estatus': a.estatus,
            })

    return render(request, 'asistencia/horas_reales.html', {
        'empleados': empleados,
        'empleado_id': int(empleado_id) if empleado_id else '',
        'desde': desde,
        'hasta': hasta,
        'resultados': resultados,
        'total': len(resultados),
    })


@login_required
def reporte_horas_secretaria(request):
    empleados = Empleado.objects.filter(estatus='activo').order_by('nombre')
    resultados = []
    empleado_id = request.GET.get('empleado_id', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')

    if empleado_id and desde and hasta:
        try:
            desde_date = date.fromisoformat(desde)
            hasta_date = date.fromisoformat(hasta)
        except (ValueError, TypeError):
            messages.error(request, 'Formato de fecha inválido. Use YYYY-MM-DD.')
            return render(request, 'asistencia/horas_secretaria.html', {
                'empleados': empleados, 'resultados': [], 'total': 0,
                'empleado_id': empleado_id, 'desde': desde, 'hasta': hasta,
            })

        delta = hasta_date - desde_date
        if delta.days > 31:
            messages.warning(request, 'El rango máximo es 31 días. Se mostrarán los primeros 31 días.')
            hasta_date = desde_date + timedelta(days=31)
            delta = hasta_date - desde_date
        try:
            empleado = Empleado.objects.get(id=empleado_id)
        except Empleado.DoesNotExist:
            messages.error(request, 'Empleado no encontrado.')
            return render(request, 'asistencia/horas_secretaria.html', {
                'empleados': empleados, 'resultados': [], 'total': 0,
            })

        for i in range(delta.days + 1):
            dia = desde_date + timedelta(days=i)
            horario_info, es_excepcion = obtener_horario_empleado(empleado, dia)
            if horario_info is None:
                continue

            if es_excepcion:
                entrada_time = horario_info.entrada_esperada
                salida_programada = horario_info.salida_esperada
                if entrada_time and salida_programada:
                    diff = datetime.combine(dia, salida_programada) - datetime.combine(dia, entrada_time)
                    jornada_hrs = round(diff.total_seconds() / 3600, 2)
                else:
                    jornada_hrs = 0
                entrada_real = entrada_time
                salida_fija = salida_programada
            else:
                horario = horario_info
                entrada_base = horario.ventana_entrada_inicio
                jornada_hrs = float(horario.jornada_hrs)

                asist = AsistenciaDiaria.objects.filter(empleado=empleado, fecha=dia).first()
                entrada_real = asist.entrada if asist and asist.entrada else entrada_base

                horas_float = jornada_hrs
                h = int(horas_float)
                m = int((horas_float - h) * 60)
                salida_dt = datetime.combine(dia, entrada_real) + timedelta(hours=h, minutes=m)
                salida_fija = salida_dt.time()

            resultados.append({
                'fecha': dia,
                'entrada': entrada_real,
                'salida_fija': salida_fija,
                'jornada': jornada_hrs,
            })

    return render(request, 'asistencia/horas_secretaria.html', {
        'empleados': empleados,
        'empleado_id': int(empleado_id) if empleado_id else '',
        'desde': desde,
        'hasta': hasta,
        'resultados': resultados,
        'total': len(resultados),
    })


@login_required
def api_recalcular(request, empleado_pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    empleado = get_object_or_404(Empleado, pk=empleado_pk)
    fecha = request.GET.get('fecha', str(timezone.localtime().date()))
    try:
        anio, mes, dia = map(int, fecha.split('-'))
        fecha_date = date(anio, mes, dia)
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Fecha inválida'}, status=400)

    asistencia = recalcular_asistencia(empleado, fecha_date)
    if asistencia:
        return JsonResponse({
            'ok': True,
            'empleado': empleado.id_original,
            'fecha': fecha,
            'estatus': asistencia.estatus,
        })
    return JsonResponse({'ok': True, 'mensaje': 'Sin datos para recalcular'})


@login_required
def api_recalcular_todos(request):
    from .calculators.engine import recalcular_todos_pendientes
    total = recalcular_todos_pendientes()
    return JsonResponse({
        'ok': True,
        'total': total,
        'mensaje': f'Recalculados {total} empleados para hoy',
    })
